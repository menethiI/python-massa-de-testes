import openpyxl
import json
from faker import Faker
from funcoesParceiro import *

# main(projName, TelaCadastro, quantidadeCadastros, contmaticID)

planilhaParceiro = openpyxl.load_workbook('excel/planilhas/importar-parceiro.xlsx')
paginaParceiro = planilhaParceiro['importar-parceiros']

planilhaPRD = openpyxl.load_workbook('excel/planilhas/importar-produto.xlsx')
paginaPRD = planilhaPRD['importar-produtos']

with open("script-py/mapeamento_cadastros.json", encoding='utf-8') as meu_json:
    dadosJson = json.load(meu_json)

camposParceiros = dadosJson["simplifique"]["clientes"]["campos"]
camposPRD = dadosJson["simplifique"]["produtos"]["campos"]


def preencherPlanilhaCamposObrigatorios(argQuantidadeCadastro):
    for linhaPlanilha in range (POSICAO_PRIMEIRA_LINHA_PREENCHIMENTO, calcQuantidadeCadastro(argQuantidadeCadastro) ):
        infDadoCriado = {}
        for colunaPlanilha, campo in [(y + 1, campo) for (y, campo) in enumerate(camposParceiros)]:
            if campo["obrigatorio"]:
                paginaParceiro.cell(row = linhaPlanilha, column = colunaPlanilha , value = checkValor(infDadoCriado, campo['algoritmo'], campo['cabecalho']))
            else:
                paginaParceiro.cell(row = linhaPlanilha, column = colunaPlanilha, value = "")

def preencherPlanilhaTodosCampos(argQuantidadeCadastro): 
    for linhaPlanilha in range (POSICAO_PRIMEIRA_LINHA_PREENCHIMENTO, calcQuantidadeCadastro(argQuantidadeCadastro) ):
        infDadoCriado = {}
        for colunaPlanilha, campo in [(y + 1, campo) for (y, campo) in enumerate(camposParceiros)]:
            paginaParceiro.cell(row = linhaPlanilha, column = colunaPlanilha , value = checkValor(infDadoCriado, campo['algoritmo'],  campo['cabecalho']))

preencherPlanilhaTodosCampos(5)
planilhaParceiro.save('excel/planilhasGerada/importar-parceiros-teste.xlsx')