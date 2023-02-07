import openpyxl
from faker import Faker
from funcoes_produto import *
import json

locales = 'pt-BR'
fake = Faker(locales)

with open("mapeamento_cadastros.json", 'r', encoding='utf-8') as mapeamento_json:
    dados = json.load(mapeamento_json)

def gerarlistaCadastroCamposObrigatorios():
    listaCadastroObrigatorio = [ '00', "Ativo", generate_codigoProduto(), generate_nomeProduto(), generate_codigoBarras(), generate_precoVenda(), generate_marca(), generate_peso(), generate_unidadeMedida()]
    return listaCadastroObrigatorio

planilha = openpyxl.load_workbook('planilhas/importar-produto.xlsx')

cadastro_parceiros_page = planilha['importar-produtos']

for x in range(3,5):
    cadastro = gerarlistaCadastroCamposObrigatorios()
    campos = dados["simplifique"]["produtos"]["campos"]
    for y, campo in [(y + 1, campo) for (y, campo) in enumerate(campos)]:         
        if campo["obrigatorio"]:
            cadastro_parceiros_page.cell(row = x, column = y, value = cadastro[y])
        else:
            cadastro_parceiros_page.cell(row = x, column = y, value = "")

planilha.save('planilhasTeste/importar-produtos-teste.xlsx')
