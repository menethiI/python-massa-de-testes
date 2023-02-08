import openpyxl
from faker import Faker
from funcoes_parceiros import *
import json

locales = 'pt-BR'
fake = Faker(locales)

with open("mapeamento_cadastros.json", 'r', encoding='utf-8') as mapeamento_json:
    dados = json.load(mapeamento_json)

def gerarlistaCadastroCamposObrigatorios():
    listaCadastroObrigatorio = [ '00', generate_status(), generate_tipoCadastro(), generate_perfilParceiro(), generate_documentoIdentificacao(), "razao social", "nome fantasia", "data", "cnae", "inscricao", generate_tipoInscricaoEstadual(), "teste"]
    return listaCadastroObrigatorio

planilha = openpyxl.load_workbook('planilhas/importar-parceiro.xlsx')

cadastro_parceiros_page = planilha['importar-parceiros']

for x in range(3,4):
    cadastro = gerarlistaCadastroCamposObrigatorios()
    campos = dados["simplifique"]["clientes"]["campos"]
    for y, campo in [(y + 1, campo) for (y, campo) in enumerate(campos)]:         
        if campo["obrigatorio"]:
            cadastro_parceiros_page.cell(row = x, column = y, value = cadastro[y])
        else:
            cadastro_parceiros_page.cell(row = x, column = y, value = "")
planilha.save('planilhasTeste/importar-parceiros-teste.xlsx')
